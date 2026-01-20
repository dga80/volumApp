from openpyxl import load_workbook

wb = load_workbook('Libro1.xlsx')
ws = wb.active
print(f'Nombre de la hoja: {ws.title}')
print(f'\nDatos (primeras 10 filas):')
for i, row in enumerate(ws.iter_rows(values_only=True, max_row=10), 1):
    print(f'Fila {i}: {row}')
